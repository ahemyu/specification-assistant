"""FastAPI service for PDF text extraction, question answering and key extraction."""
import asyncio
import json
import logging
import os
import uuid
from concurrent.futures import ProcessPoolExecutor
from io import BytesIO
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from llm_key_extractor import LLMKeyExtractor
from models import (
    ExcelDownloadRequest,
    ExcelTemplateExtractionRequest,
    ExcelTemplateResponse,
    KeyExtractionRequest,
    KeyExtractionResult,
    MultipleKeysExtractionRequest,
    QuestionRequest,
)
from openpyxl import load_workbook
from process_pdfs import process_single_pdf

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

app = FastAPI(title="PDF Text Extraction API", version="1.0.0")

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# In-memory storage for processed PDF data, for now this is file but will need to be moved to a cache
# Key: file_id, Value: processed pdf_data dict from process_single_pdf_to_dict
pdf_storage: dict[str, dict] = {}

# In-memory storage for uploaded Excel templates
# Key: template_id, Value: dict with excel_data, keys, and filled_excel (after extraction)
excel_template_storage: dict[str, dict] = {}

# Initialize LLM key extractor
llm_extractor: LLMKeyExtractor | None = None
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
if GOOGLE_API_KEY:
    try:
        llm_extractor = LLMKeyExtractor(api_key=GOOGLE_API_KEY)
        logger.info("LLM key extractor initialized successfully")
    except Exception as e:
        logger.warning(f"Failed to initialize LLM key extractor: {str(e)}")
else:
    logger.warning("GOOGLE_API_KEY not found. LLM key extraction endpoints will not be available.")


def _process_single_file(file_contents: bytes, filename: str, output_dir: Path) -> dict:
    """
    Process a single PDF file synchronously.

    This function is designed to be run in parallel via ProcessPoolExecutor.

    Args:
        file_contents: The PDF file contents as bytes
        filename: The name of the file
        output_dir: Directory to save the output text file

    Returns:
        Dictionary with processing result or error information
    """
    try:
        logger.info(f"Processing {filename}...")

        # Process PDF from memory
        pdf_data = process_single_pdf(BytesIO(file_contents), filename=filename)

        # Convert structured data to formatted text
        text_content = pdf_data["formatted_text"]

        # Save extracted text to output directory
        safe_filename = filename.replace('.pdf', '.txt')
        output_path = output_dir / safe_filename

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text_content)

        file_id = output_path.stem

        logger.info(f"Successfully processed {filename}")

        return {
            "success": True,
            "filename": filename,
            "file_id": file_id,
            "pdf_data": pdf_data
        }

    except Exception as e:
        logger.error(f"Error processing {filename}: {str(e)}")
        return {
            "success": False,
            "filename": filename,
            "error": str(e)
        }


@app.get("/")
async def root(request: Request):
    """Serve the main HTML page."""
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
async def upload_pdfs(files: list[UploadFile] = File(...)):
    """
    Upload and process multiple PDF files in-memory with parallel processing.

    Returns JSON with extracted content and file IDs for download.
    """
    processed = []
    failed = []

    # Filter out non-PDF files and read all file contents
    valid_files = []
    for file in files:
        if not file.filename or not file.filename.endswith('.pdf'):
            failed.append(f"{file.filename or 'Unknown'} (not a PDF)")
            continue

        contents = await file.read()
        valid_files.append((contents, file.filename))

    if not valid_files:
        return {"processed": processed, "failed": failed}

    # Process PDFs in parallel using ProcessPoolExecutor
    loop = asyncio.get_event_loop()
    max_workers = os.cpu_count() or 4
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # Submit all processing tasks
        tasks = [
            loop.run_in_executor(
                executor,
                _process_single_file,
                file_contents,
                filename,
                OUTPUT_DIR
            )
            for file_contents, filename in valid_files
        ]

        # Wait for all tasks to complete
        results = await asyncio.gather(*tasks)

    # Process results
    for result in results:
        if result["success"]:
            file_id = result["file_id"]
            pdf_data = result["pdf_data"]

            # Store processed PDF data in memory for later extraction
            pdf_storage[file_id] = pdf_data

            processed.append({
                "filename": result["filename"],
                "file_id": file_id,
                "total_pages": pdf_data["total_pages"],
                "data": pdf_data
            })
        else:
            failed.append(f"{result['filename']} ({result['error']})")

    return {
        "processed": processed,
        "failed": failed
    }


@app.get("/download/{file_id}")
async def download_file(file_id: str):
    """Download the extracted text file."""
    # Try with .txt extension
    file_path = OUTPUT_DIR / f"{file_id}.txt"

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        media_type="text/plain",
        filename=f"extracted_{file_id}.txt"
    )


@app.get("/preview/{file_id}")
async def preview_file(file_id: str):
    """
    Get the text content of an extracted file for preview.

    Returns JSON with the text content and metadata.
    """
    file_path = OUTPUT_DIR / f"{file_id}.txt"

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        return {
            "file_id": file_id,
            "filename": f"{file_id}.txt",
            "content": content,
            "size": len(content)
        }
    except Exception as e:
        logger.error(f"Error reading file {file_id}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )



#TODO: why do we have two functions for the same fucking thing????????
@app.post("/extract-key")
async def extract_key(request: KeyExtractionRequest) -> KeyExtractionResult:
    """
    Extract a specific key from one or more previously uploaded PDFs using LLM.

    Requires:
    - file_ids: List of file IDs from previous /upload requests
    - key_name: The key to extract (e.g., "voltage rating", "manufacturer name")
    - additional_context (optional): Additional context to help the LLM

    Returns:
    - KeyExtractionResult with the extracted value and source locations
    """
    if not llm_extractor:
        raise HTTPException(
            status_code=503,
            detail="LLM key extraction service is not available. GOOGLE_API_KEY may not be configured."
        )

    # Load the processed PDF data for each file_id from memory
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )

        pdf_data_list.append(pdf_storage[file_id])

    # Extract the key using LLM
    try:
        result = llm_extractor.extract_key(
            key_name=request.key_name,
            pdf_data=pdf_data_list,
            additional_context=request.additional_context or ""
        )
        return result
    except Exception as e:
        logger.error(f"Error during LLM key extraction: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error during key extraction: {str(e)}"
        )


@app.post("/extract-multiple-keys")
async def extract_multiple_keys(request: MultipleKeysExtractionRequest) -> dict:
    """
    Extract multiple keys from one or more previously uploaded PDFs using LLM.

    Requires:
    - file_ids: List of file IDs from previous /upload requests
    - key_names: List of keys to extract
    - additional_context (optional): Additional context to help the LLM

    Returns:
    - Dictionary mapping each key name to its KeyExtractionResult
    """
    if not llm_extractor:
        raise HTTPException(
            status_code=503,
            detail="LLM key extraction service is not available. GOOGLE_API_KEY may not be configured."
        )

    # Load the processed PDF data for each file_id from memory
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )

        pdf_data_list.append(pdf_storage[file_id])

    # Extract all keys using LLM
    try:
        results = llm_extractor.extract_multiple_keys(
            key_names=request.key_names,
            pdf_data=pdf_data_list,
            additional_context=request.additional_context or ""
        )
        # Convert results to dict with serializable values
        return {
            key: result.model_dump() if result else None
            for key, result in results.items()
        }
    except Exception as e:
        logger.error(f"Error during LLM multiple key extraction: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error during key extraction: {str(e)}"
        )


@app.post("/ask-question")
async def ask_question(request: QuestionRequest) -> dict:
    """
    Ask a general question about one or more previously uploaded PDFs using LLM.

    Requires:
    - file_ids: List of file IDs from previous /upload requests
    - question: The question to ask about the documents

    Returns:
    - Dictionary with the question and answer
    """
    if not llm_extractor:
        raise HTTPException(
            status_code=503,
            detail="LLM service is not available. GOOGLE_API_KEY may not be configured."
        )

    # Load the processed PDF data for each file_id from memory
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )

        pdf_data_list.append(pdf_storage[file_id])

    # Get answer from LLM
    try:
        # Convert conversation history to dict format for LLM
        conversation_history = None
        if request.conversation_history:
            conversation_history = [
                {"role": msg.role, "content": msg.content}
                for msg in request.conversation_history
            ]

        answer, system_message = llm_extractor.answer_question(
            question=request.question,
            pdf_data=pdf_data_list,
            conversation_history=conversation_history,
            model_name=request.model_name
        )
        response = {
            "question": request.question,
            "answer": answer,
            "document_count": len(pdf_data_list)
        }
        if system_message:
            response["system_message"] = system_message
        return response
    except Exception as e:
        logger.error(f"Error answering question: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error answering question: {str(e)}"
        )


@app.post("/ask-question-stream")
async def ask_question_stream(request: QuestionRequest):
    """
    Ask a general question about one or more previously uploaded PDFs using LLM with streaming.

    Requires:
    - file_ids: List of file IDs from previous /upload requests
    - question: The question to ask about the documents

    Returns:
    - Streaming response with Server-Sent Events (SSE) format
    """
    if not llm_extractor:
        raise HTTPException(
            status_code=503,
            detail="LLM service is not available. GOOGLE_API_KEY may not be configured."
        )
    # Load the processed PDF data for each file_id from memory
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )

        pdf_data_list.append(pdf_storage[file_id])

    # Convert conversation history to dict format for LLM
    conversation_history = None
    if request.conversation_history:
        conversation_history = [
            {"role": msg.role, "content": msg.content}
            for msg in request.conversation_history
        ]

    async def event_generator():
        """Generate SSE events for streaming response."""
        try:
            async for chunk, system_message in llm_extractor.answer_question_stream(
                question=request.question,
                pdf_data=pdf_data_list,
                conversation_history=conversation_history,
                model_name=request.model_name
            ):
                # Send system message if this is the first message
                if system_message:
                    system_event = {
                        "type": "system_message",
                        "content": system_message
                    }
                    yield f"data: {json.dumps(system_event)}\n\n"

                # Send the chunk
                chunk_event = {
                    "type": "chunk",
                    "content": chunk
                }
                yield f"data: {json.dumps(chunk_event)}\n\n"

            # Send completion event
            done_event = {
                "type": "done"
            }
            yield f"data: {json.dumps(done_event)}\n\n"

        except Exception as e:
            logger.error(f"Error during streaming: {str(e)}")
            error_event = {
                "type": "error",
                "content": str(e)
            }
            yield f"data: {json.dumps(error_event)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )




@app.post("/upload-excel-template")
async def upload_excel_template(file: UploadFile = File(...)) -> ExcelTemplateResponse:
    """
    Upload an Excel template file with 'Key' and 'Value' columns.

    The 'Key' column should be filled with keys to extract.
    The 'Value' column should be empty (will be filled by LLM).

    Returns:
    - template_id: Unique ID for this template
    - keys: List of keys found in the Key column
    - total_keys: Total number of keys
    """
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )

    try:
        # Read Excel file contents
        contents = await file.read()
        excel_file = BytesIO(contents)

        # Load workbook with openpyxl
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Validate required columns
        headers = [cell.value for cell in sheet[1]]

        if "Key" not in headers or "Value" not in headers:
            raise HTTPException(
                status_code=400,
                detail="Excel file must contain 'Key' and 'Value' columns in the first row"
            )

        # Find column indices
        key_col_index = headers.index("Key") + 1  # openpyxl is 1-indexed

        # Extract keys from Key column (skip header row)
        keys = []
        for row in sheet.iter_rows(min_row=2, min_col=key_col_index, max_col=key_col_index):
            cell_value = row[0].value
            if cell_value and str(cell_value).strip():
                keys.append(str(cell_value).strip())

        if not keys:
            raise HTTPException(
                status_code=400,
                detail="No keys found in the 'Key' column. Please fill in at least one key."
            )

        # Generate unique template ID
        template_id = str(uuid.uuid4())

        # Store the Excel file in memory for later use
        excel_file.seek(0)
        excel_template_storage[template_id] = {
            "excel_data": excel_file.read(),
            "keys": keys,
            "filename": file.filename
        }

        logger.info(f"Uploaded Excel template {template_id} with {len(keys)} keys")

        return ExcelTemplateResponse(
            template_id=template_id,
            keys=keys,
            total_keys=len(keys)
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing Excel template: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error processing Excel template: {str(e)}"
        )


@app.post("/extract-keys-from-template")
async def extract_keys_from_template(request: ExcelTemplateExtractionRequest) -> dict:
    """
    Extract keys from PDFs using an uploaded Excel template.

    Requires:
    - template_id: ID from previous /upload-excel-template request
    - file_ids: List of file IDs from /upload requests
    - additional_context (optional): Additional context for extraction

    Returns:
    - Dictionary of extraction results (same format as /extract-multiple-keys)
    """
    if not llm_extractor:
        raise HTTPException(
            status_code=503,
            detail="LLM service is not available. GOOGLE_API_KEY may not be configured."
        )

    # Verify template exists
    if request.template_id not in excel_template_storage:
        raise HTTPException(
            status_code=404,
            detail=f"Template with ID {request.template_id} not found. Please upload the template first."
        )

    # Load the processed PDF data for each file_id
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )
        pdf_data_list.append(pdf_storage[file_id])

    # Get template data
    template_data = excel_template_storage[request.template_id]
    keys = template_data["keys"]
    excel_bytes = template_data["excel_data"]
    filename = template_data["filename"]

    # Extract all keys using LLM
    try:
        results = llm_extractor.extract_multiple_keys(
            key_names=keys,
            pdf_data=pdf_data_list,
            additional_context=request.additional_context or ""
        )

        # Load Excel template with openpyxl to preserve formatting
        excel_file = BytesIO(excel_bytes)
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Find column indices
        headers = [cell.value for cell in sheet[1]]
        value_col_index = headers.index("Value") + 1

        # Fill in the Value column
        current_row = 2  # Start after header
        for key in keys:
            result = results.get(key)
            if result:
                value = result.key_value if result.key_value else "Not found"
            else:
                value = "Extraction failed"

            # Write value to the Value column
            sheet.cell(row=current_row, column=value_col_index, value=value)
            current_row += 1

        # Save filled Excel to storage for later download
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Store filled Excel in template storage
        excel_template_storage[request.template_id]["filled_excel"] = output.read()
        excel_template_storage[request.template_id]["filled_filename"] = f"filled_{filename}"

        # Return extraction results as JSON (same format as /extract-multiple-keys)
        return {
            key: result.model_dump() if result else None
            for key, result in results.items()
        }

    except Exception as e:
        logger.error(f"Error extracting keys from template: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error extracting keys: {str(e)}"
        )


@app.get("/download-filled-excel/{template_id}")
async def download_filled_excel(template_id: str):
    """
    Download the filled Excel template after extraction.

    Args:
        template_id: The template ID from the extraction

    Returns:
        Filled Excel file with preserved formatting
    """
    if template_id not in excel_template_storage:
        raise HTTPException(
            status_code=404,
            detail="Template not found. Please upload and extract first."
        )

    template_data = excel_template_storage[template_id]

    if "filled_excel" not in template_data:
        raise HTTPException(
            status_code=404,
            detail="Filled Excel not found. Please extract keys first."
        )

    filled_excel_bytes = template_data["filled_excel"]
    filename = template_data.get("filled_filename", "filled_template.xlsx")

    return StreamingResponse(
        BytesIO(filled_excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.post("/download-extraction-excel")
async def download_extraction_excel(request: ExcelDownloadRequest):
    """
    Download extraction results as an Excel file with two columns: Key and Value.

    Requires:
    - extraction_results: Dictionary mapping key names to their extraction results

    Returns:
    - Excel file (.xlsx) with extracted key-value pairs
    """
    try:
        # Prepare data for Excel
        data_rows = []

        for key_name, result in request.extraction_results.items():
            # Handle None results (failed extractions)
            if result is None:
                data_rows.append({
                    "Key": key_name,
                    "Value": "Extraction failed"
                })
                continue

            # Extract value from result
            # Handle both single key results and multiple key results
            if isinstance(result, dict):
                key_value = result.get("key_value", "Not found")
            else:
                key_value = getattr(result, "key_value", "Not found")

            data_rows.append({
                "Key": key_name,
                "Value": key_value if key_value is not None else "Not found"
            })

        # Create DataFrame
        df = pd.DataFrame(data_rows)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Extracted Keys')

        output.seek(0)

        # Return as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=extracted_keys.xlsx"
            }
        )

    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel file: {str(e)}"
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

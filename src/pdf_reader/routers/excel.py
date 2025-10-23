"""Router for Excel template endpoints."""
import logging
import uuid
from io import BytesIO

import pandas as pd
from dependencies import get_excel_storage, get_llm_extractor, get_pdf_storage
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from schemas.requests import ExcelDownloadRequest, ExcelTemplateExtractionRequest
from schemas.responses import ExcelTemplateResponse
from services.llm_key_extractor import LLMKeyExtractor

logger = logging.getLogger(__name__)

router = APIRouter(prefix="", tags=["excel"])


@router.post("/upload-excel-template")
async def upload_excel_template(file: UploadFile = File(...)) -> ExcelTemplateResponse:
    """
    Upload an Excel template file with 'Key' and 'Value' columns.

    The 'Key' column should be filled with keys to extract.
    The 'Value' column should be empty (will be filled by LLM).

    Returns:
    - template_id: Unique ID for this template
    - keys: List of keys found in the Key column
    - total_keys: Total number of keys
    """
    excel_storage = get_excel_storage()

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )

    try:
        # Read Excel file contents
        contents = await file.read()
        excel_file = BytesIO(contents)

        # Load workbook with openpyxl
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Validate required columns
        headers = [cell.value for cell in sheet[1]]

        if "Key" not in headers or "Value" not in headers:
            raise HTTPException(
                status_code=400,
                detail="Excel file must contain 'Key' and 'Value' columns in the first row"
            )

        # Find column indices
        key_col_index = headers.index("Key") + 1  # openpyxl is 1-indexed

        # Extract keys from Key column (skip header row)
        keys = []
        for row in sheet.iter_rows(min_row=2, min_col=key_col_index, max_col=key_col_index):
            cell_value = row[0].value
            if cell_value and str(cell_value).strip():
                keys.append(str(cell_value).strip())

        if not keys:
            raise HTTPException(
                status_code=400,
                detail="No keys found in the 'Key' column. Please fill in at least one key."
            )

        # Generate unique template ID
        template_id = str(uuid.uuid4())

        # Store the Excel file in memory for later use
        excel_file.seek(0)
        excel_storage[template_id] = {
            "excel_data": excel_file.read(),
            "keys": keys,
            "filename": file.filename
        }

        logger.info(f"Uploaded Excel template {template_id} with {len(keys)} keys")

        return ExcelTemplateResponse(
            template_id=template_id,
            keys=keys,
            total_keys=len(keys)
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing Excel template: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error processing Excel template: {str(e)}"
        )


@router.post("/extract-keys-from-template")
async def extract_keys_from_template(
    request: ExcelTemplateExtractionRequest,
    llm_extractor: LLMKeyExtractor = Depends(get_llm_extractor)
) -> dict:
    """
    Extract keys from PDFs using an uploaded Excel template.

    Requires:
    - template_id: ID from previous /upload-excel-template request
    - file_ids: List of file IDs from /upload requests
    - additional_context (optional): Additional context for extraction

    Returns:
    - Dictionary of extraction results (same format as /extract-keys)
    """
    pdf_storage = get_pdf_storage()
    excel_storage = get_excel_storage()

    # Verify template exists
    if request.template_id not in excel_storage:
        raise HTTPException(
            status_code=404,
            detail=f"Template with ID {request.template_id} not found. Please upload the template first."
        )

    # Load the processed PDF data for each file_id
    pdf_data_list = []
    for file_id in request.file_ids:
        if file_id not in pdf_storage:
            raise HTTPException(
                status_code=404,
                detail=f"File with ID {file_id} not found. Please upload the file first."
            )
        pdf_data_list.append(pdf_storage[file_id])

    # Get template data
    template_data = excel_storage[request.template_id]
    keys = template_data["keys"]
    excel_bytes = template_data["excel_data"]
    filename = template_data["filename"]

    # Extract all keys using LLM
    try:
        results = llm_extractor.extract_keys(
            key_names=keys,
            pdf_data=pdf_data_list,
            additional_context=request.additional_context or ""
        )

        # Load Excel template with openpyxl to preserve formatting
        excel_file = BytesIO(excel_bytes)
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Find column indices
        headers = [cell.value for cell in sheet[1]]
        value_col_index = headers.index("Value") + 1

        # Fill in the Value column
        current_row = 2  # Start after header
        for key in keys:
            result = results.get(key)
            if result:
                value = result.key_value if result.key_value else "Not found"
            else:
                value = "Extraction failed"

            # Write value to the Value column
            sheet.cell(row=current_row, column=value_col_index, value=value)
            current_row += 1

        # Save filled Excel to storage for later download
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Store filled Excel in template storage
        excel_storage[request.template_id]["filled_excel"] = output.read()
        excel_storage[request.template_id]["filled_filename"] = f"filled_{filename}"

        # Return extraction results as JSON (same format as /extract-multiple-keys)
        return {
            key: result.model_dump() if result else None
            for key, result in results.items()
        }

    except Exception as e:
        logger.error(f"Error extracting keys from template: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error extracting keys: {str(e)}"
        )


@router.get("/download-filled-excel/{template_id}")
async def download_filled_excel(template_id: str):
    """
    Download the filled Excel template after extraction.

    Args:
        template_id: The template ID from the extraction

    Returns:
        Filled Excel file with preserved formatting
    """
    excel_storage = get_excel_storage()

    if template_id not in excel_storage:
        raise HTTPException(
            status_code=404,
            detail="Template not found. Please upload and extract first."
        )

    template_data = excel_storage[template_id]

    if "filled_excel" not in template_data:
        raise HTTPException(
            status_code=404,
            detail="Filled Excel not found. Please extract keys first."
        )

    filled_excel_bytes = template_data["filled_excel"]
    filename = template_data.get("filled_filename", "filled_template.xlsx")

    return StreamingResponse(
        BytesIO(filled_excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.post("/download-extraction-excel")
async def download_extraction_excel(request: ExcelDownloadRequest):
    """
    Download extraction results as an Excel file with two columns: Key and Value.

    Requires:
    - extraction_results: Dictionary mapping key names to their extraction results

    Returns:
    - Excel file (.xlsx) with extracted key-value pairs
    """
    try:
        # Prepare data for Excel
        data_rows = []

        for key_name, result in request.extraction_results.items():
            # Handle None results (failed extractions)
            if result is None:
                data_rows.append({
                    "Key": key_name,
                    "Value": "Extraction failed"
                })
                continue

            # Extract value from result
            # Handle both single key results and multiple key results
            if isinstance(result, dict):
                key_value = result.get("key_value", "Not found")
            else:
                key_value = getattr(result, "key_value", "Not found")

            data_rows.append({
                "Key": key_name,
                "Value": key_value if key_value is not None else "Not found"
            })

        # Create DataFrame
        df = pd.DataFrame(data_rows)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Extracted Keys')

        output.seek(0)

        # Return as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=extracted_keys.xlsx"
            }
        )

    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel file: {str(e)}"
        )
